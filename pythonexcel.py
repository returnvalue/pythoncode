# Excel Data Frame
import pandas as pd
from openpyxl.workbook import Workbook

dataframe_excel = pd.read_excel('excel/regiondata.xlsx')

print(dataframe_excel)

# CSV Data Frame
dataframe_csv = pd.read_csv('excel/some_names.csv')

print(dataframe_csv)

# Setting Column Names
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']
print(dataframe_csv)

# Pandas CSV to Excel
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']
print(dataframe_csv)

dataframe_csv.to_excel('excel/some_names_modified.xlsx')

# Text File Data Frame
dataframe_txt = pd.read_csv('excel/some_data.txt', delimiter='\t')

print(dataframe_txt)

# Work With Columns In Pandas
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']
print(dataframe_csv.columns)


# take the First Name index and this prints out all the values in the First Name column
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']
print(dataframe_csv['First'])

#  access multiple column’s data
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']
print(dataframe_csv[['Address', 'State']])

# look at the Zip column, but only the first two results
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']
print(dataframe_csv['Zip'][0:2])

# Work With Rows In Pandas
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']
print(dataframe_csv.iloc[2])

# Locate exact cell
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']
print(dataframe_csv.iloc[3, 2])

# Saving Extracted Data
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']

extracted_data = dataframe_csv[['First', 'Last', 'City']]

stored = extracted_data.to_excel('extracted_data.xlsx', index=None)

# Filter And Sort Data Using Pandas
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']

print(dataframe_csv[dataframe_csv['City'] == 'Worthington'])

# checks for all rows where the City is Kentwood *and* the First column has a value of Sam.
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']

print(dataframe_csv[(dataframe_csv['City'] == 'Kentwood') & (dataframe_csv['First'] == 'Sam')])

# drop columns using the .drop() function.
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']

drop = ['Address', 'Population']
dataframe_csv.drop(columns=drop, inplace=True)

print(dataframe_csv)

# check to see if the State column has a value of OH, and if it does, go ahead and set the new column we defined to True.
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']

drop = ['Address', 'Population']
dataframe_csv.drop(columns=drop, inplace=True)

dataframe_csv['T or F'] = False
dataframe_csv.loc[dataframe_csv['State'] == 'OH', 'T or F'] = True

print(dataframe_csv)

# use the .sort_values() method to sort the data on a particular column
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']

print(dataframe_csv.sort_values('First'))

# To sort the data in the other direction, just add ascending=False as the second argument
dataframe_csv = pd.read_csv('excel/some_names.csv', header=None)

dataframe_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Zip', 'Population']

print(dataframe_csv.sort_values('First', ascending=False))

# Controlling Excel Directly With Openpyxl
import openpyxl

workbook = openpyxl.load_workbook('excel/stock_options.xlsx')

print(type(workbook))

# How To Access Worksheets
workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheet = workbook['Sheet1']

print(type(sheet))

# check what names exist with a simple call to .sheetnames.
workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheetnames = workbook.sheetnames

print(sheetnames)

# Access Cells In Sheets
workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheet = workbook['Sheet1']
cell = sheet['A3']

print(cell.value)

# access a cell using the .cell() method and passing both the row and column as integers
workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheet = workbook['Sheet1']
cell = sheet.cell(row=4, column=14)

print(cell.value)

# iterate over values in the sheet
workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheet = workbook['Sheet1']

for i in range(2, 7):
    cell = sheet.cell(row=i, column=1)
    print(cell.value)

# use slicing to select a range of cells
workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheet = workbook['Sheet1']

cell_range = sheet['A1':'A3']

print(cell_range)

# print out the number of items in a column
workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheet = workbook['Sheet1']

column_a = sheet['A']

print(len(column_a))

# show all of the cells that have values in row 1
workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheet = workbook['Sheet1']

row_0 = sheet[1]

print(row_0)

# iterating through columns or rows
workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheet = workbook['Sheet1']

for row in sheet.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

# Creating New Workbooks and Worksheets
workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet2 = workbook.create_sheet('First Sheet')
worksheet3 = workbook.create_sheet('Second Sheet')

worksheet.title = 'My Awesome Sheet'

print(workbook.sheetnames)

# add some data to one of the Worksheets
workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet2 = workbook.create_sheet('First Sheet')
worksheet3 = workbook.create_sheet('Second Sheet')

worksheet.title = 'My Awesome Sheet'
worksheet['A1'] = 'Hello Openpyxl'
workbook.save('excel/awesomeworkbook.xlsx')

# How To Format Workbooks
from openpyxl.styles import Font, Alignment, GradientFill

workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheet = workbook['Sheet1']
sheet.insert_rows(1, 2)
sheet.merge_cells('A1:O2')
cell = sheet['A1']
cell.font = Font(color='007742', size=20, italic=True)
cell.value = 'Super Cool And Stylish Spreadsheet'
cell.alignment = Alignment(horizontal='right', vertical='center')
cell.fill = GradientFill(stop=('000000', 'ffffff'))
workbook.save('excel/stylish.xlsx')

# Named Styles In Openpyxl
from openpyxl.styles import Font, Alignment, GradientFill, NamedStyle, Side, Border, PatternFill

workbook = openpyxl.load_workbook('excel/stock_options.xlsx')
sheet = workbook['Sheet1']
sheet.insert_rows(1, 2)
sheet.merge_cells('A1:O2')
cell = sheet['A1']
cell.font = Font(color='007742', size=20, italic=True)
cell.value = 'Super Cool And Stylish Spreadsheet'
cell.alignment = Alignment(horizontal='right', vertical='center')
cell.fill = GradientFill(stop=('000000', 'ffffff'))

highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True)
bd = Side(style='thick', color='000000')
highlight.border = Border(left=None, top=bd, right=None, bottom=bd)
highlight.fill = PatternFill('solid', fgColor='fde295')

for cell in sheet['3:3']:
    cell.style = highlight

workbook.save('excel/stylish.xlsx')

